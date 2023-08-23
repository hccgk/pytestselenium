"""
Created by HE on 2023/8/23.

This is a description of the file.
"""
# 读取excel里面的数据
import openpyxl


def read_excel():
    workbook = openpyxl.load_workbook('../data.xlsx')
    sheet = workbook['serlogin']
    print(sheet.max_row, sheet.max_column)
    listdict = []
    for row in range(2, sheet.max_row + 1):
        # tempList = []
        empty_dict = {}
        for col in range(1, sheet.max_column + 1):
            empty_dict[sheet.cell(1, col).value] = sheet.cell(row, col).value
            # tempList.append(sheet.cell(row,col).value)
        print(empty_dict)
        listdict.append(empty_dict)
    return listdict

