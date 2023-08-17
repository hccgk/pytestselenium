import unittest

import openpyxl
from ddt import ddt,data,unpack


def read_excel():
    print('path')
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook['login']
    print(sheet.max_row,sheet.max_column)
    allList = []
    for row in range(2,sheet.max_row+1):
        tempList = []
        for col in range(2,sheet.max_column+1):
            tempList.append(sheet.cell(row,col).value)
        print(tempList)
        allList.append(tempList)
    return allList


@ddt
class Test(unittest.TestCase):

    @data({'1', '2', '3'})
    def test_01_login(self,args):
        print(args)

    @data(('1', '2'),('4','5'))
    def test_02_login(self,args):
        print(args)

    @data(*read_excel())
    @unpack
    def test_03_login(self,args1,args2):
        print(args1,args2)





if __name__ == '__main__':
    unittest.main()
    # read_excel()